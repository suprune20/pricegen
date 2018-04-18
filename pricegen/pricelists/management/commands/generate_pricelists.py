# ! /usr/bin/env python3
#
# generate_pricelists.py
#

# Параметр: корневой каталог, внутри которого кладутся pricelists
# и где формируются результаты
#
#   -   Пусть корневой каталог pricege
#   -   имеются организация-продавец korona, у нее точки выдачи
#       * korona pickpoints: zamok, kurasovschina
#   -   поставщики krynka, myasokombinat
#
# Тогда каталог будет выглядеть:
#
# ...pricegen/                                                         (1)
#           korona/
#               suppliers/                                          (2)
#                   /krynka/
#                       some_name1.xlsx (входящий файл)
#                       some_name2.xlsx (входящий файл)
#                   /myasokombinat/
#                       some_name3.xlsx (входящий файл)
#                       some_name4.xlsx (входящий файл)
#               outbox/
#                   korona_zamok_retail_ГГГГММДДЧЧММ.xlsx
#                   korona_zamok_wholesale_ГГГГММДДЧЧММ.xlsx
#                   korona_kurasovschina_retail_ГГГГММДДЧЧММ.xlsx
#                   korona_kurasovschina_wholesale_ГГГГММДДЧЧММ.xlsx
#               inbox_archive/ (упакованные копии вх.файлов)
#                   krynka_ГГГГММДДЧЧММСС.xlsx.gz
#                   krynka_ГГГГММДДЧЧММСС.xlsx.gz
#                   myasokombinat_ГГГГММДДЧЧММСС.xlsx.gz
#                   myasokombinat_ГГГГММДДЧЧММСС.xlsx.gz
#           ...
#           log/
#               pricegen-stat-ГГГГММДД.log (текущий)
#               pricegen-stat-ГГГГММДД.log.gz (за предыдущие даты)
#               pricegen-err-ГГГГММДД.log (текущий)
#               pricegen-err-ГГГГММДД.log.gz (за предыдущие даты)

import time, os, pwd, re, datetime, shutil
import subprocess

from openpyxl import load_workbook, Workbook
from openpyxl.styles import Font, Alignment

from django.conf import settings
from django.core.management.base import BaseCommand
from django.db.models import Min

from pricegen.utils import time_human, round_decimal

from users.models import Org, PickPoint
from pricelists.models import ExcelFormat, ExcelTempo, PickPointDelivery, PickPointBrand, \
    Marge

class Command(BaseCommand):
    help = "generate pricelists after input files detected"

    root_folder = None
    # Дескрипторы журналов
    #
    fd = dict(
        stat=None,
        error=None,
        log=None,
    )

    def add_arguments(self, parser):
        parser.add_argument('root_folder')

    def handle(self, *args, **options):
        root_folder = options['root_folder']
        self.root_folder = root_folder

        # Нельзя запускать, если другой аналогичные процесс крутится
        #
        ps = subprocess.check_output(('ps', '-eo', 'user:30,pid:10,cmd')).decode('utf-8').split('\n')
        userid = pwd.getpwuid(os.getuid()).pw_name
        pid = os.getpid()
        for p in ps:
            # пропускаем процессы с другим именем
            #
            if not re.search(r'manage\.py\s+generate_pricelists\s+%s' % re.escape(root_folder), p):
                continue
            # пропускаем наш процесс: с нашим именем пользователя и pid процесса
            #
            if re.search(r'^%s\s+%s\s+' % (userid, pid,), p):
                continue
            print('Another same process running. Do not want problems. Quit.')
            quit()

        # Создать каталог для журналов, если он не существует
        #
        quarantine_folder = os.path.join(root_folder, settings.FS_QUARANTINE_FOLDER)
        log_folder = os.path.join(root_folder, settings.FS_LOG_FOLDER)
        tmp_folder = os.path.join(root_folder, settings.FS_TMP_FOLDER)
        archive_folder = os.path.join(root_folder, settings.FS_ARCHIVE_FOLDER)
        for folder in (quarantine_folder, log_folder, tmp_folder, archive_folder,):
            if not os.path.isdir(folder):
                try:
                    os.mkdir(folder)
                except OSError:
                    print('ERROR: Failed to create %s folder' % folder)
                    quit()
        for tmp_file in os.listdir(tmp_folder):
            path_to_tmp_file = os.path.join(tmp_folder, tmp_file)
            if os.path.isfile(path_to_tmp_file):
                os.unlink(path_to_tmp_file)

        self.write_log('Started generate pricelists', 'log')

        # Глобальный цикл. Выходим из него, когда не найдем файлов для обработки
        #
        while True:
            found_input = False

            # Цикл (1). Читаем по всем организациям, все они могут быть продавцами
            #
            self.write_log("Started search in vendors' folders", 'log')
            for vendor in Org.objects.all():
                
                # Если у продавца нет pickpoints или не найден supplier
                # по имени папки suppliers/<supplier>, то входные 
                #
                ignore_vendor_input = False
                vendor_folder = os.path.join(root_folder, settings.FS_VENDORS_FOLDER, vendor.short_name)
                
                # Цикл (2) . По каталогам, имена которых организации-поставщики
                #
                try:
                    vendors_suppliers_folder = os.path.join(vendor_folder, settings.FS_SUPPLIERS_FOLDER)
                    suppliers_folders = os.listdir(vendors_suppliers_folder)
                except FileNotFoundError:
                    # Например, нет каталога ...pricegen/krynka/suppliers.
                    # Если krynka - только поставщик, но никак не покупатель, то ОК
                    #
                    continue
                outbox_folder = os.path.join(vendor_folder, settings.FS_VENDOR_OUTBOX_FOLDER)
                if not os.path.isdir(outbox_folder):
                    os.mkdir(outbox_folder)

                # Есть ли продавца pickpoints. Заодно запомним pickpoints
                #
                pickpoint_deliveries_to = PickPointDelivery.objects.filter(pickpoint_to__org=vendor). \
                    select_related('pickpoint_from')

                if not pickpoint_deliveries_to:
                    self.write_log("No pickpoint deliveries at vendor organization: %s. Move all vendor's input, if any, to quarantine" % (
                            vendor.short_name
                        ), 'error')
                    ignore_vendor_input = True
                else:
                    output_col_numbers = self.xlsx_col_numbers(vendor, input_=False)
                    output_sheet_rows = 0
                    for item in output_col_numbers:
                        output_sheet_rows = max(output_sheet_rows, output_col_numbers[item])
                    output_sheet_rows += 1

                for supplier_folder in suppliers_folders:
                    ignore_supplier_input = False
                    path_to_supplier_folder = os.path.join(vendors_suppliers_folder, supplier_folder)
                    if not os.path.isdir(path_to_supplier_folder):
                        continue
                    try:
                        supplier = Org.objects.get(short_name=supplier_folder)
                    except Org.DoesNotExist:
                        ignore_supplier_input = True
                        self.write_log("%s, no appropriate organization: %s.  Move all input to the supplier, if any, to quarantine" % (
                                path_to_supplier_folder,
                                supplier_folder
                            ), 'error')
                    xlsx_files = []
                    for xlsx_file in os.listdir(path_to_supplier_folder):
                        path_to_xlsx_file = os.path.join(path_to_supplier_folder, xlsx_file)
                        if not os.path.isfile(path_to_xlsx_file) or \
                               os.path.islink(path_to_xlsx_file) or \
                           not re.search(r'\.xlsx$', xlsx_file, flags=re.I):
                            continue
                        if ignore_vendor_input or ignore_supplier_input:
                            self.put_to_quarantine(path_to_xlsx_file)
                            continue
                        stat = os.stat(path_to_xlsx_file)
                        xlsx_files.append(dict(
                            name=xlsx_file,
                            mtime=stat.st_mtime
                        ))
                    if ignore_vendor_input or ignore_supplier_input:
                        continue
                    xlsx_files = sorted(xlsx_files, key=lambda d: d['mtime'])
                    for xlsx_file_dict in xlsx_files:
                        xlsx_file = xlsx_file_dict['name']
                        self.write_log('Found input "%s", supplier %s to vendor %s' % (
                                xlsx_file,
                                supplier_folder,
                                vendor.short_name,
                            ),'log')
                        found_input = True
                        path_to_xlsx_file = os.path.join(path_to_supplier_folder, xlsx_file)
                        if not self.load_xlsx_to_tempo(path_to_xlsx_file, supplier):
                            self.write_log('Error reading ExcelX file: %s' % (
                                    path_to_xlsx_file,
                                ), 'error')
                            self.put_to_quarantine(path_to_xlsx_file)
                            continue
                        pickpoints_to = list()
                        is_smth_to_xlsx = False
                        for pickpoint_delivery_to in pickpoint_deliveries_to:
                            pickpoint_to = pickpoint_delivery_to.pickpoint_to
                            # Будут ли какие-то записи в
                            #   vendor_pickpoint_retail_ГГГГММДДЧЧММ.xlsx
                            #   vendor_pickpoint_wholesale_ГГГГММДДЧЧММ.xlsx
                            #
                            are_quantities_int = True
                            marges_retail = self.get_marges('retail', pickpoint_to)
                            marges_wholesale = self.get_marges('wholesale', pickpoint_to)
                            for pickpoint_to_brand in PickPointBrand.objects.filter(pickpoint=pickpoint_to):
                                # Здесь вычислить минимальное время доставки этого brand
                                # к этому pickpoint_to, mvd
                                #
                                mvd = PickPointDelivery.objects.filter(
                                    pickpoint_to__org=vendor,
                                    pickpoint_from__pickpointbrand__brand=pickpoint_to_brand.brand,
                                ).aggregate(Min('delivery_time'))
                                mvd = mvd['delivery_time__min']
                                if mvd is None:
                                    # Нет среди складов поставщиков такого, чтоб там был
                                    # этот brand
                                    continue
                                mvd_human = time_human(mvd)
                                pickpoint_to_brand_name = pickpoint_to_brand.brand.name
                                for xlsx_rec in ExcelTempo.objects.filter(
                                   brand__iexact=pickpoint_to_brand_name,
                                   ).order_by('row'):
                                    if not is_smth_to_xlsx:
                                        output_book_retail = Workbook()
                                        output_sheet_retail = output_book_retail.active
                                        output_sheet_retail.title = settings.XLSX_OUTPUT_SHEET_NAME
                                        output_book_wholesale = Workbook()
                                        output_sheet_wholesale = output_book_wholesale.active
                                        output_sheet_wholesale.title = settings.XLSX_OUTPUT_SHEET_NAME
                                        n_rows = 0
                                        is_smth_to_xlsx = True
                                    output_row = ['' for i in range(output_sheet_rows)]
                                    output_row[output_col_numbers['inner_id_col']] = xlsx_rec.inner_id
                                    output_row[output_col_numbers['partnumber_col']] = xlsx_rec.partnumber
                                    output_row[output_col_numbers['brand_col']] = xlsx_rec.brand
                                    output_row[output_col_numbers['item_name_col']] = xlsx_rec.item_name
                                    output_row[output_col_numbers['quantity_col']] = str(xlsx_rec.quantity)
                                    if are_quantities_int and xlsx_rec.quantity - int(xlsx_rec.quantity):
                                        are_quantities_int = False
                                    output_row[output_col_numbers['delivery_time_col']] = mvd_human

                                    output_row[output_col_numbers['price_col']] = \
                                        str(self.apply_marge(xlsx_rec.price, marges_retail))
                                    output_sheet_retail.append(output_row)

                                    output_row[output_col_numbers['price_col']] = \
                                        str(self.apply_marge(xlsx_rec.price, marges_wholesale))
                                    output_sheet_wholesale.append(output_row)
                                    n_rows += 1

                            self.xlsx_format(output_sheet_retail, output_col_numbers, n_rows, are_quantities_int)
                            self.xlsx_format(output_sheet_wholesale, output_col_numbers, n_rows, are_quantities_int)
                            now_ = datetime.datetime.now().strftime('%Y%m%d%H%M%S')
                            parts = dict(
                                vendor=vendor.short_name,
                                pickpoint=pickpoint_to.short_name,
                                now_=now_,
                            )
                            xlsx_retail_name = self.check_same_fname(
                                outbox_folder,
                                "%(vendor)s_%(pickpoint)s_retail_%(now_)s.xlsx" % parts,
                            )
                            xlsx_wholesale_name = self.check_same_fname(
                                outbox_folder,
                                "%(vendor)s_%(pickpoint)s_wholesale_%(now_)s.xlsx" % parts,
                            )
                            tmp_xlsx_retail = os.path.join(tmp_folder, xlsx_retail_name)
                            tmp_xlsx_wholesale = os.path.join(tmp_folder, xlsx_wholesale_name)
                            output_book_retail.save(tmp_xlsx_retail)
                            output_book_wholesale.save(tmp_xlsx_wholesale)
                            shutil.move(tmp_xlsx_retail, outbox_folder)
                            shutil.move(tmp_xlsx_wholesale, outbox_folder)
                            self.write_log('Result %s is in %s outbox folder' % (
                                    xlsx_retail_name,
                                    vendor.short_name,
                                ),'log')
                            self.write_log('Result %s is in %s outbox folder' % (
                                    xlsx_wholesale_name,
                                    vendor.short_name,
                                ),'log')
                        dt_xlsx_input = datetime.datetime.fromtimestamp(xlsx_file_dict['mtime'])
                        dt_xlsx_input = dt_xlsx_input.strftime('%Y%m%d%H%M%S')
                        xlsx_archive_file = '%s~%s_%s.xlsx' % (
                            vendor.short_name,
                            supplier.short_name,
                            dt_xlsx_input
                        )
                        path_to_xlsx_archive_file = os.path.join(archive_folder, xlsx_archive_file) 
                        try:
                            shutil.move(path_to_xlsx_file, path_to_xlsx_archive_file)
                        except shutil.Error:
                            # Может быть, если такой файл в архиве уже есть
                            # А вдруг раньше недо-переместили в архив?
                            # Если это не учесть, может быть бесконечный цикл
                            #
                            os.unlink(path_to_xlsx_archive_file)
                            shutil.move(path_to_xlsx_file, path_to_xlsx_archive_file)
                        self.zip_(path_to_xlsx_archive_file)
                        self.write_log('%s input %s archived as %s.%s' % (
                                vendor.short_name,
                                xlsx_file,
                                xlsx_archive_file,
                                settings.ZIP_EXT,
                            ),'log')

            self.write_log("End of search in vendors' folders", 'log')
            if not found_input:
                break
        self.write_log("No files found in vendors' folders. Quit.", 'log')

    def check_same_fname(self, folder, fname):
        """
        Проверить, есть ли такое fname в folder, и если есть изменить fname

        Имя изменяется так: fname.ext -> fname-1.ext
        """
        if not os.path.isfile(os.path.join(folder, fname)):
            return fname
        m = re.search(r'^(.+)\.([^\.]+)$', fname)
        if m:
            fname_ = m.group(1)
            point_ = '.'
            ext_ = m.group(2)
        else:
            fname_ = fname_
            point_ = ''
            ext_ = ''
        add_ = 1
        while True:
            result = '%s-%s%s%s' % (
                    fname_,
                    str(add_),
                    point_,
                    ext_,
            )
            if os.path.isfile(os.path.join(folder, result)):
                add_ += 1
            else:
                break
        return result

    def apply_marge(self, num, marges):
        """
        Применить маржу с округлением
        """
        result = num
        if marges:
            for marge in marges:
                if num < marge['limit']:
                    break
                else:
                    result = num * (1 + marge['marge'] / 100)
        return round_decimal(result)

    def xlsx_format(self, output_sheet, output_col_numbers, n_rows, are_quantities_int):
        """
        Форматирование выходного xlsx файла
        """

        # Список соответствий
        # 0-я колонка Excel- файла : 'brand_col'
        # 1-я колонка Excel- файла : 'item_name_col'
        #
        map_output = dict()
        for item in output_col_numbers:
            map_output[output_col_numbers[item]] = item
        for i, column_cells in enumerate(output_sheet.columns):
            try:
                item = map_output[i]
            except KeyError:
                continue
            width = settings.XLSX_COL_STYLES[item]['width']
            font = Font(**settings.XLSX_COL_STYLES[item]['font'])
            alignment = Alignment(**settings.XLSX_COL_STYLES[item]['alignment'])
            output_sheet.column_dimensions[column_cells[0].column].width = width
            
            # А это по всей колонке не задашь! Для LibreOffice так можно,
            # а вот чтоб Microsoft Excel воспринимал, надо устанавливать
            # font, alignment по каждой ячейке из колонки
            #
            for n_row in range(n_rows):
                if are_quantities_int and item == 'quantity_col':
                    column_cells[n_row].value = int(float(column_cells[n_row].value))
                column_cells[n_row].font = font
                column_cells[n_row].alignment = alignment


    def get_marges(self, kind, pickpoint_to):
        """
        Получить маржи по пункту продажи, сортированные по limit
        """
        marges = list()
        for marge in Marge.objects.filter(pickpoint=pickpoint_to, kind=kind). \
            order_by('pickpoint', 'limit'):
            marges.append(dict(
                limit=marge.limit,
                marge=marge.marge,
            ))
        return marges

    def put_to_quarantine(self, path_to_xlsx_file):
        utc_time = int(time.time())
        dt = datetime.datetime.fromtimestamp(utc_time)
        dt_folder = dt.strftime('%Y-%m-%d~%H-%M-%S')
        dt_folder = os.path.join(self.root_folder, settings.FS_QUARANTINE_FOLDER, dt_folder)
        os.mkdir(dt_folder)
        shutil.move(path_to_xlsx_file, dt_folder)
        self.write_log('%s moved to %s' % (
                path_to_xlsx_file,
                dt_folder,
            ), 'error')

    def load_xlsx_to_tempo(self, path_to_xlsx_file, supplier):
        """
        Загрузить Excel файл во временную таблицу
        
        Параметры:
            path_to_xlsx_file
            supplier: организация, у нее ищем формат Excel файла
        """
        input_col_numbers = self.xlsx_col_numbers(supplier, input_=True)
        try:
            wb = load_workbook(filename=path_to_xlsx_file)
        except:
            return False
        sheet = wb.active
        ExcelTempo.objects.all().delete()
        rows = sheet.rows
        nrow = 0
        for row in rows:
            nrow += 1
            rec = ExcelTempo()
            rec.row = nrow
            for field in input_col_numbers:
                # a_field_col -> a_field
                db_field = field[:-4]
                # print (field, input_col_numbers[field], row[input_col_numbers[field]])
                setattr(rec, db_field, row[input_col_numbers[field]].value)
            rec.save()
        return True

    def xlsx_col_numbers(self, org, input_):
        """
        Получить номера колонок в Excel файле
        
        Параметр:
            vendor:     организация, у нее ищем формат Excel файла
            input_:     True для входного файла, False для выходного
        """
        try:
            xf = ExcelFormat.objects.get(org=org)
            input_col_numbers = dict(
                inner_id_col=xf.inner_id_col,
                partnumber_col=xf.partnumber_col,
                brand_col=xf.brand_col,
                item_name_col=xf.item_name_col,
                price_col=xf.price_col,
                quantity_col=xf.quantity_col,
                delivery_time_col=xf.delivery_time_col,
            )
        except ExcelFormat.DoesNotExist:
            input_col_numbers = settings.XLSX_COL_NUMBERS_DEFAULT.copy()
        if input_:
            for k in settings.XLSX_OUTPUT_ONLY_COLS:
                del input_col_numbers[k]
        for item in input_col_numbers:
            input_col_numbers[item] -=1
        return input_col_numbers

    def zip_(self, path_to_fname):
        """
        Запаковать файл
        """
        zipped_file = '%s.%s' % (path_to_fname, settings.ZIP_EXT)
        if os.path.isfile(zipped_file):
            os.unlink(zipped_file)
        os.system('%s %s' % (settings.CMD_ZIP, path_to_fname))
        return zipped_file

    def write_log(self, rec, kind='stat'):
        """
        Запись в журнал, статистики или ошибок: kind = stat или error или log
        
        Журналы ведутся по принципу, например, для статистики:
        pricegen-stat-ГГГГММДД.log (текущий)
        pricegen-stat-ГГГГММДД.log.gz (за предыдущие даты)
        """
        if kind == 'stat':
            log_name_prefix = settings.FS_STAT_PREFIX
        elif kind == 'log':
            log_name_prefix = settings.FS_LOG_PREFIX
        elif kind == 'error':
            log_name_prefix = settings.FS_ERROR_PREFIX
        else:
            return
        utc_time = int(time.time())
        dt = datetime.datetime.fromtimestamp(utc_time)
        yyyymmdd = '%d%02d%02d' % (dt.year, dt.month, dt.day)
        log_name = '%s%s.%s' % (log_name_prefix, yyyymmdd, settings.FS_LOG_EXT)
        log_folder = os.path.join(self.root_folder, settings.FS_LOG_FOLDER)
        log_name = os.path.join(log_folder, log_name)
        if not os.path.isfile(log_name):
            # Заархивировать все предыдущие журналы
            #
            for f in self.fd:
                if self.fd[f]:
                    self.fd[f].close()
                    self.fd[f] = None
            for f in os.listdir(log_folder):
                fname = os.path.join(log_folder, f)
                if os.path.isfile(fname) and \
                   re.search(r'^%s\d{8}\.%s$' % (
                       re.escape(log_name_prefix),
                       re.escape(settings.FS_LOG_EXT),
                   ), f):
                    self.zip_(fname)
        if not self.fd[kind]:
            self.fd[kind] = open(log_name, 'a')
        if kind == 'stat':
            dt_prefix = '%s\t' % utc_time
        else:
            dt_prefix = dt.strftime('%Y-%m-%d %H:%M:%S ')
        self.fd[kind].write('%s%s\n' % (dt_prefix, rec,))
