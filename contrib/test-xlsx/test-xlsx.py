#! /usr/bin/env python3

# test read/write xlsx
#
# Моделируем чтение xlsx файла и запись в него

# Запуск из ./manage.py shell:
#   exec(open('/path/to/test-xlsx.py').read())

# Колонки, ширины c пробного прайс-листа, начиная с 1
# при шрифте Arial 9
#
# 1 11.6640625      (inner_id_col)
# 2 17.5            (partnumber_col)
# 3 18.6640625      (brand_col)
# 4 88.1640625      (item_name_col)
# 5 11.33203125     (price_col)
# 6 10.6640625      (quantity_col)

INPUT_FILE = '/home/suprune20/musor/test-xlsx/zzap_p30.xlsx'
OUTPUT_FILE = '/home/suprune20/musor/test-xlsx/output.xlsx'

# Так в этих INPUT_FILE, OUTPUT_FILES. В других может быть иначе.
# Как надо, берется из базы или из settings.
#
input_col_numbers = dict(
    inner_id_col=1,
    partnumber_col=2,
    brand_col=3,
    item_name_col=4,
    price_col=5,
    quantity_col=6,
)
output_col_numbers = dict(
    inner_id_col=3,
    partnumber_col=4,
    brand_col=1,
    item_name_col=2,
    price_col=7,
    quantity_col=6,
    delivery_time_col=5
)

# Это будет вычисляться. А здесь затычка (plug)
#
DELIVERY_TYME_PLUG = 130

# Для установки в settings.py
# ---------------------------

OUTPUT_SHEET_NAME = 'TDSheet'

# Параметры колонок.
#
COL_STYLES = dict(
    inner_id_col=dict(
        width=12,
        font=dict(
            name='Arial',
            size=9,
        ),
        alignment=dict(
            horizontal='left'
        ),
    ),
    partnumber_col=dict(
        width=18,
        font=dict(
            name='Arial',
            size=9,
        ),
        alignment=dict(
            horizontal='left'
        ),
    ),
    brand_col=dict(
        width=19,
        font=dict(
            name='Arial',
            size=9,
        ),
        alignment=dict(
            horizontal='left'
        ),
    ),
    item_name_col=dict(
        width=90,
        font=dict(
            name='Arial',
            size=9,
            bold=True,
        ),
        alignment=dict(
            horizontal='left'
        ),
    ),
    price_col=dict(
        width=12,
        font=dict(
            name='Arial',
            size=9,
        ),
        alignment=dict(
            horizontal='right'
        ),
    ),
    quantity_col=dict(
        width=10,
        font=dict(
            name='Arial',
            size=9,
        ),
        alignment=dict(
            horizontal='right'
        ),
    ),
    delivery_time_col=dict(
        width=12,
        font=dict(
            name='Arial',
            size=9,
        ),
        alignment=dict(
            horizontal='left'
        ),
    ),
)

# -----------

from openpyxl import load_workbook, Workbook
from openpyxl.styles import Font, Alignment

def main():
    input_wb = load_workbook(filename = INPUT_FILE)
    input_sheet = input_wb.active

    output_book = Workbook()
    output_sheet = output_book.active
    output_sheet.title = OUTPUT_SHEET_NAME

    # В базе данных номера колонок начинаются с 1.
    # Приведем эти номера к "машинному виду, т.е. с нуля
    #
    # Заодно получим число колонок в файлах:
    #
    input_sheet_rows = 0
    output_sheet_rows = 0
    
    for item in input_col_numbers:
        input_sheet_rows = max(input_sheet_rows, input_col_numbers[item])
        input_col_numbers[item] -=1
    for item in output_col_numbers:
        output_sheet_rows = max(output_sheet_rows, output_col_numbers[item])
        output_col_numbers[item] -=1
    
    # Теперь список соответствий
    # 0 (inner_id_col in input) : 2 (inner_id_col in output)
    # 1 (partnumber_col in input) : 3 (partnumber_col in output)
    # ...
    #
    map_col_number = [None for i in range(output_sheet_rows)]
    for item in input_col_numbers:
        map_col_number[input_col_numbers[item]] = output_col_numbers[item]

    rows = input_sheet.rows
    for row in rows:
        input_row = ['' for i in range(max(input_sheet_rows ,output_sheet_rows))]
        for i, cell in enumerate(row):
            input_row[i] = cell.value
        output_row = ['' for i in range(output_sheet_rows)]
        output_row[output_col_numbers['delivery_time_col']] = time_human(DELIVERY_TYME_PLUG)
        for i_input, i_output in enumerate(map_col_number):
            if i_output is None:
                # delivery_time_col, уже занесли
                pass
            else:
                output_row[i_output] = input_row[i_input]
        output_sheet.append(output_row)

    #for i, column_cells in enumerate(input_sheet.columns):
        #length = input_sheet.column_dimensions[column_cells[0].column].width
        #output_sheet.column_dimensions[column_cells[0].column].width = length

    # Список соответствий
    # 0-я колонка Excel- файла : 'brand_col'
    # 1-я колонка Excel- файла : 'item_name_col'
    #
    map_output = dict()
    for item in output_col_numbers:
        map_output[output_col_numbers[item]] = item
    for i, column_cells in enumerate(output_sheet.columns):
        try:
            item = map_output[i]
        except KeyError:
            continue
        width = COL_STYLES[item]['width']
        font = Font(**COL_STYLES[item]['font'])
        alignment = Alignment(**COL_STYLES[item]['alignment'])
        output_sheet.column_dimensions[column_cells[0].column].width = width
        output_sheet.column_dimensions[column_cells[0].column].font = font
        output_sheet.column_dimensions[column_cells[0].column].alignment = alignment

    output_book.save(OUTPUT_FILE)

def time_human(mins):
    """
    Время из минут в '?? час. ?? мин.'
    """
    mins_ = mins % 60
    hours_ = int(mins/60)
    result = ''
    if hours_:
        result += '%s час.' % hours_
    if mins_ and hours_:
        result += ' '
    if mins_:
        result += '%s мин.' % mins_
    return result

main()
